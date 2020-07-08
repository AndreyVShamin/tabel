from openpyxl import load_workbook, formula
from openpyxl.styles import Border, Side
import sqlite3
import calendar


"""
index
eventday
fullname
tabnum
rank
startdate
stopdate
sickleave
vacation
holiday
"""


def mergeDate():
    for col in range (4, 38):
        sheet.merge_cells(start_row=9, start_column=col, end_row=11, end_column=col)


def newEmployee( n, summative ):
    firstRow = n*2 + 11
    for col in range ( 1, 4 ):
        sheet.merge_cells( start_row=firstRow, start_column=col, end_row=firstRow + 1, end_column=col )
        
    for col in range ( 36, 42 ):
        sheet.merge_cells( start_row=firstRow, start_column=col, end_row=firstRow + 1, end_column=col )
        
    sheet.merge_cells( start_row=firstRow, start_column=45, end_row=firstRow + 1, end_column=45 )
    #sheet.cell( column = 19, row = firstRow, value = "=SUM(RC[-15]:RC[-1]")
    #sheet.cell( column = 36, row = firstRow, value = "=SUM(RC[-15]:RC[-1]")
    
    
    for col in range ( 1, 46 ):
        sheet.cell( row=firstRow, column=col ).border = Border (top=thin, left=thin, right=thin, bottom=thin)
        sheet.cell( row=firstRow + 1, column=col ).border = Border (top=thin, left=thin, right=thin, bottom=thin)

    days = 0
    hoursFull = 0
    for day in range ( 1, 15+1 ):
        col = day + 3
        dayOfWeek = calendar.weekday(year, month, day)
        if dayOfWeek == 5 or dayOfWeek == 6:
            val = 'в'
        else:
            days += 1
            hoursFull += 8
    sheet.cell( column = 19, row = firstRow, value = hoursFull )
    summative[0] += hoursFull    
        
    hoursHalf = 0
    for day in range ( 16, calendar.monthrange(year, month)[1]+1 ):
        col = day + 4
        dayOfWeek = calendar.weekday( year, month, day )
        if dayOfWeek == 5 or dayOfWeek == 6:
            val = 'в'
        else:
            days += 1
            hoursFull += 8
            hoursHalf += 8
    sheet.cell( column = 36, row = firstRow, value = hoursHalf )
    summative[1] += hoursHalf
    summative[2] += days
    summative[3] += hoursFull
        
    sheet.cell( column = 37, row = firstRow, value = days )
    sheet.cell( column = 38, row = firstRow, value = hoursFull )
    
    return summative

        
def footer( footerFirstRow ):
    fr = footerFirstRow
    sheet.cell( column = 3, row = fr).border = Border( bottom=thin )
    sheet.cell( column = 7, row = fr).border = Border( bottom=thin )
    sheet.cell( column = 12, row = fr).border = Border( bottom=thin )
    sheet.merge_cells( start_row=fr, start_column=3, end_row=fr, end_column=5 )
    sheet.merge_cells( start_row=fr, start_column=7, end_row=fr, end_column=10 )
    sheet.merge_cells( start_row=fr, start_column=12, end_row=fr, end_column=16 )
    sheet.cell( column = 2, row = fr, value = 'Ответственное лицо' )
    sheet.cell( column = 3, row = fr, value = 'зам. директора' )
    sheet.cell( column = 12, row = fr, value = 'Шамин А.В.' )
    fr = footerFirstRow + 1
    sheet.cell( column = 29, row = fr).border = Border( bottom=thin )
    sheet.cell( column = 35, row = fr).border = Border( bottom=thin )
    sheet.cell( column = 38, row = fr).border = Border( bottom=thin )
    sheet.merge_cells( start_row=fr, start_column=3, end_row=fr, end_column=5 )
    sheet.merge_cells( start_row=fr, start_column=7, end_row=fr, end_column=10 )
    sheet.merge_cells( start_row=fr, start_column=12, end_row=fr, end_column=17 )
    sheet.merge_cells( start_row=fr, start_column=21, end_row=fr, end_column=25 )
    sheet.merge_cells( start_row=fr, start_column=29, end_row=fr, end_column=33 )
    sheet.merge_cells( start_row=fr, start_column=35, end_row=fr, end_column=37 )
    sheet.merge_cells( start_row=fr, start_column=38, end_row=fr, end_column=41 )
    sheet.merge_cells( start_row=fr, start_column=43, end_row=fr, end_column=45 )
    sheet.cell( column = 3, row = fr, value = '(должность)' )
    sheet.cell( column = 7, row = fr, value = '(личная подпись)' )
    sheet.cell( column = 12, row = fr, value = '(расшифровка подписи)' )
    sheet.cell( column = 21, row = fr, value = 'Руководитель' )
    sheet.cell( column = 29, row = fr, value = 'директор' )
    sheet.cell( column = 38, row = fr, value = 'Панова Е.А.' ) 
    sheet.cell( column = 43, row = fr, value = '"30" сентября 2017 г.' )
    fr = footerFirstRow + 2
    sheet.merge_cells( start_row=fr, start_column=29, end_row=fr, end_column=33 )
    sheet.merge_cells( start_row=fr, start_column=35, end_row=fr, end_column=37 )
    sheet.merge_cells( start_row=fr, start_column=38, end_row=fr, end_column=41 )
    sheet.cell( column = 29, row = fr, value = '(должность)' )
    sheet.cell( column = 35, row = fr, value = '(личная подпись)' )
    sheet.cell( column = 38, row = fr, value = '(расшифровка подписи)' )
    fr = footerFirstRow + 3
    sheet.cell( column = 29, row = fr).border = Border( bottom=thin )
    sheet.cell( column = 35, row = fr).border = Border( bottom=thin )
    sheet.cell( column = 38, row = fr).border = Border( bottom=thin )
    sheet.merge_cells( start_row=fr, start_column=21, end_row=fr, end_column=28 )
    sheet.merge_cells( start_row=fr, start_column=29, end_row=fr, end_column=33 )
    sheet.merge_cells( start_row=fr, start_column=35, end_row=fr, end_column=37 )
    sheet.merge_cells( start_row=fr, start_column=38, end_row=fr, end_column=41 )
    sheet.merge_cells( start_row=fr, start_column=43, end_row=fr, end_column=45 )
    sheet.cell( column = 21, row = fr, value = 'Работник кадровой службы' )
    sheet.cell( column = 29, row = fr, value = 'директор' )
    sheet.cell( column = 38, row = fr, value = 'Панова Е.А.' ) 
    sheet.cell( column = 43, row = fr, value = '"30" сентября 2017 г.' )
    fr = footerFirstRow + 4
    sheet.merge_cells( start_row=fr, start_column=29, end_row=fr, end_column=33 )
    sheet.merge_cells( start_row=fr, start_column=35, end_row=fr, end_column=37 )
    sheet.merge_cells( start_row=fr, start_column=38, end_row=fr, end_column=41 )
    sheet.cell( column = 29, row = fr, value = '(должность)' )
    sheet.cell( column = 35, row = fr, value = '(личная подпись)' )
    sheet.cell( column = 38, row = fr, value = '(расшифровка подписи)' )
    fr = footerFirstRow + 5
    sheet.cell( column = 29, row = fr).border = Border( bottom=thin )
    sheet.cell( column = 35, row = fr).border = Border( bottom=thin )
    sheet.cell( column = 38, row = fr).border = Border( bottom=thin )
    sheet.merge_cells( start_row=fr, start_column=21, end_row=fr, end_column=26 )
    sheet.merge_cells( start_row=fr, start_column=29, end_row=fr, end_column=33 )
    sheet.merge_cells( start_row=fr, start_column=35, end_row=fr, end_column=37 )
    sheet.merge_cells( start_row=fr, start_column=38, end_row=fr, end_column=41 )
    sheet.merge_cells( start_row=fr, start_column=43, end_row=fr, end_column=45 )
    sheet.cell( column = 21, row = fr, value = 'Работник ООТиЗ' )
    sheet.cell( column = 29, row = fr, value = 'директор' )
    sheet.cell( column = 38, row = fr, value = 'Панова Е.А.' ) 
    sheet.cell( column = 43, row = fr, value = '"30" сентября 2017 г.' )
    fr = footerFirstRow + 6
    sheet.merge_cells( start_row=fr, start_column=29, end_row=fr, end_column=33 )
    sheet.merge_cells( start_row=fr, start_column=35, end_row=fr, end_column=37 )
    sheet.merge_cells( start_row=fr, start_column=38, end_row=fr, end_column=41 )
    sheet.cell( column = 29, row = fr, value = '(должность)' )
    sheet.cell( column = 35, row = fr, value = '(личная подпись)' )
    sheet.cell( column = 38, row = fr, value = '(расшифровка подписи)' )    
    

conn = sqlite3.connect( 'c:/temp/tabel.db3' )
wb = load_workbook( 'c:/Temp/Табель.xlsx' )

year = 2020
month = 7

sheet = wb['tab']
thin = Side(border_style="thin", color="000000")
"""
summative = [summativeI,summativeII, summativeDays, summativeHours]
"""
summative = [0,0,0,0]
summative = newEmployee( 1, summative )
print (summative)
summative = newEmployee( 2, summative )
print (summative)
summative = newEmployee( 3, summative )
print (summative)
summative = newEmployee( 4, summative )
print (summative)
summative = newEmployee( 5, summative )
print (summative)
summative = newEmployee( 6, summative )
summativeRow = 6*2 + 11 + 2
sheet.cell( column = 19, row = summativeRow, value = summative[0] )
sheet.cell( column = 36, row = summativeRow, value = summative[1] )
sheet.cell( column = 37, row = summativeRow, value = summative[2] )
sheet.cell( column = 38, row = summativeRow, value = summative[3] )

footer (26)    
wb.save('c:/Temp/ТабельNew.xlsx')
conn.close()
print('The end.')


